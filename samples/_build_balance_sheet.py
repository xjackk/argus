#!/usr/bin/env python3
"""Build LARGE, richly-formatted Balance Sheet demo workbooks (v1/v2).

A rolling balance sheet across 7 periods (FY20 opening + FY21-FY26) with 24
line items across Assets / Liabilities / Equity. It BALANCES BY CONSTRUCTION
via the accounting identity: Cash is rolled from a cash-flow build, PP&E from
capex/depreciation, Debt from draws, Retained Earnings from net income -- so
Total Assets = Total Liabilities & Equity in EVERY period, in BOTH versions.

An Assumptions sheet (blue input cells) drives everything via cross-sheet
formulas. v1 -> v2 changes ONE input (Annual Capex), funded by long-term debt
draws, so it ripples through PP&E, depreciation, cash, debt and the totals
across all forecast periods while the balance check stays $0.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CUR = '$#,##0'
PCT = '0.0%'
INPUT_FILL = PatternFill(fill_type='solid', fgColor='DCE6F1')
INPUT_FONT = Font(color='1F4E78', bold=True)
TITLE_FONT = Font(bold=True, size=13, color='1F4E78')
HDR_FONT = Font(bold=True, color='FFFFFF')
HDR_FILL = PatternFill(fill_type='solid', fgColor='1F4E78')
SECTION_FONT = Font(bold=True, color='1F4E78')
BOLD = Font(bold=True)
TOP_BORDER = Border(top=Side(style='thin', color='808080'))

PERIODS = ['FY20', 'FY21', 'FY22', 'FY23', 'FY24', 'FY25', 'FY26']
FIRST_COL = 2  # column B (FY20 = opening, static)


def build(capex, path):
    wb = Workbook()

    # ---------------- Assumptions ----------------
    a = wb.active
    a.title = 'Assumptions'
    a['A1'] = 'Balance Sheet Assumptions'
    a['A1'].font = TITLE_FONT
    a['A2'] = '(blue cells are authored inputs)'
    a['A2'].font = Font(italic=True, size=9, color='808080')

    inputs = [
        # Opening (FY20) balances
        ('Opening Cash ($000)',            18000, CUR),  # B3
        ('Opening Accounts Receivable',    9000,  CUR),  # B4
        ('Opening Inventory',              7000,  CUR),  # B5
        ('Opening Prepaid Expenses',       1500,  CUR),  # B6
        ('Opening Gross PP&E',             40000, CUR),  # B7
        ('Opening Accumulated Deprec.',    12000, CUR),  # B8
        ('Goodwill',                       15000, CUR),  # B9
        ('Intangible Assets',              6000,  CUR),  # B10
        ('Opening Accounts Payable',       6500,  CUR),  # B11
        ('Opening Accrued Expenses',       3000,  CUR),  # B12
        ('Opening Deferred Revenue',       2500,  CUR),  # B13
        ('Opening Short-term Debt',        5000,  CUR),  # B14
        ('Opening Long-term Debt',         22000, CUR),  # B15
        ('Common Stock',                   30000, CUR),  # B16
        # Annual drivers
        ('Annual Capex ($000)',            capex, CUR),  # B17  <-- THE CHANGE
        ('Depreciation Rate (of Gross)',   0.10,  PCT),  # B18
        ('Annual Net Income',              6000,  CUR),  # B19
        ('AR Growth per Year',             600,   CUR),  # B20
        ('Inventory Growth per Year',      500,   CUR),  # B21
        ('AP Growth per Year',             400,   CUR),  # B22
    ]
    AR = {}
    r = 3
    for label, val, fmt in inputs:
        a.cell(row=r, column=1, value=label)
        c = a.cell(row=r, column=2, value=val)
        c.number_format = fmt
        c.fill = INPUT_FILL
        c.font = INPUT_FONT
        AR[label] = r
        r += 1
    a.column_dimensions['A'].width = 30
    a.column_dimensions['B'].width = 14

    def A(label):
        return f"Assumptions!$B${AR[label]}"

    # ---------------- Balance Sheet ----------------
    s = wb.create_sheet('BalanceSheet')
    s['A1'] = 'Balance Sheet (USD $000)'
    s['A1'].font = TITLE_FONT
    s['A2'] = '($ in thousands)'
    s['A2'].font = Font(italic=True, size=9, color='808080')
    for i, per in enumerate(PERIODS):
        cell = s.cell(row=2, column=FIRST_COL + i, value=per)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='right')

    # rows: (key|None, label, kind) kind in {section,item,total,check}
    layout = [
        (None,    'ASSETS',                         'section'),
        ('cash',  'Cash & Equivalents',             'item'),
        ('ar',    'Accounts Receivable',            'item'),
        ('inv',   'Inventory',                      'item'),
        ('prep',  'Prepaid Expenses',               'item'),
        ('tca',   'Total Current Assets',           'total'),
        ('gppe',  'Gross PP&E',                     'item'),
        ('accd',  'Less: Accumulated Depreciation', 'item'),
        ('nppe',  'Net PP&E',                       'total'),
        ('gw',    'Goodwill',                       'item'),
        ('intan', 'Intangible Assets',              'item'),
        ('tnca',  'Total Non-Current Assets',       'total'),
        ('ta',    'Total Assets',                   'total'),
        (None,    'LIABILITIES',                    'section'),
        ('ap',    'Accounts Payable',               'item'),
        ('accr',  'Accrued Expenses',               'item'),
        ('defr',  'Deferred Revenue',               'item'),
        ('std',   'Short-term Debt',                'item'),
        ('tcl',   'Total Current Liabilities',      'total'),
        ('ltd',   'Long-term Debt',                 'item'),
        ('tl',    'Total Liabilities',              'total'),
        (None,    'EQUITY',                         'section'),
        ('cs',    'Common Stock',                   'item'),
        ('re',    'Retained Earnings',              'item'),
        ('te',    'Total Equity',                   'total'),
        ('tle',   'Total Liabilities & Equity',     'total'),
        ('chk',   'Balance Check (A - L&E)',        'check'),
    ]
    R = {}
    r = 3
    for key, label, kind in layout:
        lc = s.cell(row=r, column=1, value=label)
        if kind == 'section':
            lc.font = SECTION_FONT
        elif kind in ('total', 'check'):
            lc.font = BOLD
        if key:
            R[key] = r
        r += 1
    total_keys = {k for k, _, kind in layout if kind in ('total', 'check')}

    for i, per in enumerate(PERIODS):
        col = get_column_letter(FIRST_COL + i)
        prev = get_column_letter(FIRST_COL + i - 1) if i > 0 else None
        opening = (i == 0)

        def put(key, formula):
            c = s.cell(row=R[key], column=FIRST_COL + i, value=formula)
            c.number_format = CUR
            if key in total_keys:
                c.font = BOLD
                c.border = TOP_BORDER

        if opening:
            # FY20 opening balances straight from Assumptions
            put('cash', f"={A('Opening Cash ($000)')}")
            put('ar',   f"={A('Opening Accounts Receivable')}")
            put('inv',  f"={A('Opening Inventory')}")
            put('prep', f"={A('Opening Prepaid Expenses')}")
            put('gppe', f"={A('Opening Gross PP&E')}")
            put('accd', f"={A('Opening Accumulated Deprec.')}")
            put('gw',   f"={A('Goodwill')}")
            put('intan',f"={A('Intangible Assets')}")
            put('ap',   f"={A('Opening Accounts Payable')}")
            put('accr', f"={A('Opening Accrued Expenses')}")
            put('defr', f"={A('Opening Deferred Revenue')}")
            put('std',  f"={A('Opening Short-term Debt')}")
            put('ltd',  f"={A('Opening Long-term Debt')}")
            put('cs',   f"={A('Common Stock')}")
            # opening retained earnings is the balancing figure so the opening
            # sheet ties: RE = TotalAssets - TotalLiab - CommonStock
            put('re',   (f"=({col}{R['tca']}+{col}{R['tnca']})"
                         f"-{col}{R['tl']}-{col}{R['cs']}"))
        else:
            # roll forward. Capex funded by a long-term debt draw of equal size.
            put('cash', (f"={prev}{R['cash']}+{A('Annual Net Income')}"
                         f"+({col}{R['accd']}-{prev}{R['accd']})"          # + depreciation
                         f"-{A('Annual Capex ($000)')}"                    # - capex
                         f"+{A('Annual Capex ($000)')}"                    # + debt draw (funds capex)
                         f"-{A('AR Growth per Year')}-{A('Inventory Growth per Year')}"
                         f"+{A('AP Growth per Year')}"))
            put('ar',   f"={prev}{R['ar']}+{A('AR Growth per Year')}")
            put('inv',  f"={prev}{R['inv']}+{A('Inventory Growth per Year')}")
            put('prep', f"={prev}{R['prep']}")
            put('gppe', f"={prev}{R['gppe']}+{A('Annual Capex ($000)')}")
            put('accd', f"={prev}{R['accd']}+{col}{R['gppe']}*{A('Depreciation Rate (of Gross)')}")
            put('gw',   f"={prev}{R['gw']}")
            put('intan',f"={prev}{R['intan']}")
            put('ap',   f"={prev}{R['ap']}+{A('AP Growth per Year')}")
            put('accr', f"={prev}{R['accr']}")
            put('defr', f"={prev}{R['defr']}")
            put('std',  f"={prev}{R['std']}")
            put('ltd',  f"={prev}{R['ltd']}+{A('Annual Capex ($000)')}")   # debt draw funds capex
            put('cs',   f"={prev}{R['cs']}")
            put('re',   f"={prev}{R['re']}+{A('Annual Net Income')}")

        # subtotals / totals (same every column)
        put('tca',  f"=SUM({col}{R['cash']}:{col}{R['prep']})")
        put('nppe', f"={col}{R['gppe']}-{col}{R['accd']}")
        put('tnca', f"={col}{R['nppe']}+{col}{R['gw']}+{col}{R['intan']}")
        put('ta',   f"={col}{R['tca']}+{col}{R['tnca']}")
        put('tcl',  f"=SUM({col}{R['ap']}:{col}{R['std']})")
        put('tl',   f"={col}{R['tcl']}+{col}{R['ltd']}")
        put('te',   f"={col}{R['cs']}+{col}{R['re']}")
        put('tle',  f"={col}{R['tl']}+{col}{R['te']}")
        put('chk',  f"={col}{R['ta']}-{col}{R['tle']}")

    s.column_dimensions['A'].width = 32
    for i in range(len(PERIODS)):
        s.column_dimensions[get_column_letter(FIRST_COL + i)].width = 12

    wb.save(path)
    print('wrote', path)


if __name__ == '__main__':
    build(4000, 'balance_sheet_v1.xlsx')
    build(9000, 'balance_sheet_v2.xlsx')
