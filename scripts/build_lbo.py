from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE = Font(name='Arial', color='0000FF')       # inputs
BLACK = Font(name='Arial', color='000000')       # formulas
GREEN = Font(name='Arial', color='008000')       # cross-sheet links
BOLD = Font(name='Arial', bold=True)
BOLDW = Font(name='Arial', bold=True, color='FFFFFF')
HDR = PatternFill('solid', fgColor='1F4E5F')
SUB = PatternFill('solid', fgColor='D9E1E8')
YEL = PatternFill('solid', fgColor='FFF4D6')
thin = Side(style='thin', color='C9D1DA')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(ws, row, cols, text=None):
    for c in cols:
        cell = ws.cell(row, c)
        cell.fill = HDR; cell.font = BOLDW; cell.border = BORDER
        cell.alignment = Alignment(horizontal='center')

def build(path):
    wb = Workbook()

    # ---------- ASSUMPTIONS ----------
    a = wb.active; a.title = 'Assumptions'
    a.column_dimensions['A'].width = 34
    for col in 'BCDEF': a.column_dimensions[col].width = 13
    a['A1'] = 'Project Atlas — LBO Assumptions'; a['A1'].font = Font(name='Arial', bold=True, size=14)
    rows = [
        (3, 'Transaction', None, True),
        (4, 'Entry EV / EBITDA', 9.0, 'x'),
        (5, 'Exit EV / EBITDA', 10.5, 'x'),
        (6, 'Entry Net Debt / EBITDA', 4.5, 'x'),
        (7, 'Hold Period (years)', 5, 'int'),
        (9, 'Operating', None, True),
        (10, 'Entry Revenue ($mm)', 500.0, '$'),
        (11, 'Revenue Growth (% p.a.)', 0.08, '%'),
        (12, 'Entry EBITDA Margin (%)', 0.22, '%'),
        (13, 'Exit EBITDA Margin (%)', 0.25, '%'),
        (15, 'Financing', None, True),
        (16, 'Interest Rate on Debt (%)', 0.075, '%'),
        (17, 'Mandatory Debt Paydown (% p.a.)', 0.10, '%'),
        (18, 'Tax Rate (%)', 0.25, '%'),
    ]
    for r, label, val, kind in rows:
        a.cell(r, 1, label)
        if kind is True:
            a.cell(r, 1).font = BOLD; a.cell(r, 1).fill = SUB
            for c in range(2, 6): a.cell(r, c).fill = SUB
        else:
            cell = a.cell(r, 2, val); cell.font = BLUE; cell.fill = YEL; cell.border = BORDER
            if kind == '%': cell.number_format = '0.0%'
            elif kind == 'x': cell.number_format = '0.0x'
            elif kind == '$': cell.number_format = '$#,##0'
            else: cell.number_format = '0'

    # ---------- P&L (projection, formula-driven off Assumptions) ----------
    p = wb.create_sheet('P&L')
    p.column_dimensions['A'].width = 30
    for i, col in enumerate('BCDEFG'): p.column_dimensions[col].width = 13
    p['A1'] = 'P&L Projection ($mm)'; p['A1'].font = Font(name='Arial', bold=True, size=14)
    p['A3'] = 'Year'; p['A3'].font = BOLD
    for yr in range(0, 6):
        c = p.cell(3, 2+yr, f'Y{yr}'); c.font = BOLDW; c.fill = HDR; c.alignment = Alignment(horizontal='center')

    # Revenue: Y0 = entry revenue; Yn = prior*(1+growth)
    p['A4'] = 'Revenue'
    p['B4'] = "=Assumptions!B10"; p['B4'].font = GREEN
    for yr in range(1, 6):
        col = get_column_letter(2+yr); prev = get_column_letter(1+yr)
        p[f'{col}4'] = f"={prev}4*(1+Assumptions!$B$11)"; p[f'{col}4'].font = BLACK
    # EBITDA margin ramps linearly entry->exit across hold
    p['A5'] = 'EBITDA Margin'
    for yr in range(0, 6):
        col = get_column_letter(2+yr)
        # linear interpolation entry margin -> exit margin over hold period
        p[f'{col}5'] = f"=Assumptions!$B$12+(Assumptions!$B$13-Assumptions!$B$12)*{yr}/Assumptions!$B$7"
        p[f'{col}5'].font = BLACK; p[f'{col}5'].number_format = '0.0%'
    # EBITDA = Revenue * margin
    p['A6'] = 'EBITDA'; p['A6'].font = BOLD
    for yr in range(0, 6):
        col = get_column_letter(2+yr)
        p[f'{col}6'] = f"={col}4*{col}5"; p[f'{col}6'].font = BLACK; p[f'{col}6'].font = BOLD
    # D&A assume 4% of revenue
    p['A7'] = 'Less: D&A (4% rev)'
    for yr in range(0, 6):
        col = get_column_letter(2+yr)
        p[f'{col}7'] = f"=-{col}4*0.04"; p[f'{col}7'].font = BLACK
    # EBIT
    p['A8'] = 'EBIT'
    for yr in range(0, 6):
        col = get_column_letter(2+yr)
        p[f'{col}8'] = f"={col}6+{col}7"; p[f'{col}8'].font = BLACK
    # Interest expense links to Debt schedule opening balance
    p['A9'] = 'Less: Interest'; p['A9'].font = GREEN
    for yr in range(0, 6):
        col = get_column_letter(2+yr)
        if yr == 0:
            p[f'{col}9'] = 0
        else:
            p[f'{col}9'] = f"=-Debt!{col}4*Assumptions!$B$16"; p[f'{col}9'].font = GREEN
    # EBT, tax, NI
    p['A10'] = 'EBT'
    for yr in range(0, 6):
        col = get_column_letter(2+yr)
        p[f'{col}10'] = f"={col}8+{col}9"; p[f'{col}10'].font = BLACK
    p['A11'] = 'Less: Tax'
    for yr in range(0, 6):
        col = get_column_letter(2+yr)
        p[f'{col}11'] = f"=-MAX({col}10,0)*Assumptions!$B$18"; p[f'{col}11'].font = BLACK
    p['A12'] = 'Net Income'; p['A12'].font = BOLD
    for yr in range(0, 6):
        col = get_column_letter(2+yr)
        p[f'{col}12'] = f"={col}10+{col}11"; p[f'{col}12'].font = BOLD
    for r in range(4, 13):
        for yr in range(0, 6):
            col = get_column_letter(2+yr)
            if p[f'{col}{r}'].number_format == 'General':
                p[f'{col}{r}'].number_format = '$#,##0;($#,##0);-'

    # ---------- DEBT SCHEDULE ----------
    d = wb.create_sheet('Debt')
    d.column_dimensions['A'].width = 30
    for col in 'BCDEFG': d.column_dimensions[col].width = 13
    d['A1'] = 'Debt Schedule ($mm)'; d['A1'].font = Font(name='Arial', bold=True, size=14)
    d['A3'] = 'Year'; d['A3'].font = BOLD
    for yr in range(0, 6):
        c = d.cell(3, 2+yr, f'Y{yr}'); c.font = BOLDW; c.fill = HDR; c.alignment = Alignment(horizontal='center')
    # Opening debt = entry net debt/EBITDA * entry EBITDA
    d['A4'] = 'Opening Debt'; d['A4'].font = GREEN
    d['B4'] = "=Assumptions!B6*'P&L'!B6"; d['B4'].font = GREEN
    for yr in range(1, 6):
        col = get_column_letter(2+yr); prev = get_column_letter(1+yr)
        d[f'{col}4'] = f"={prev}5"; d[f'{col}4'].font = BLACK   # opening = prior closing
    # Mandatory paydown = -opening * paydown%
    d['A5'] = 'Closing Debt'; d['A5'].font = BOLD
    for yr in range(0, 6):
        col = get_column_letter(2+yr)
        d[f'{col}5'] = f"={col}4*(1-Assumptions!$B$17)"; d[f'{col}5'].font = BLACK; d[f'{col}5'].font = BOLD
    for r in (4, 5):
        for yr in range(0, 6):
            col = get_column_letter(2+yr)
            d[f'{col}{r}'].number_format = '$#,##0;($#,##0);-'

    # ---------- RETURNS ----------
    r = wb.create_sheet('Returns')
    r.column_dimensions['A'].width = 34
    r.column_dimensions['B'].width = 16
    r['A1'] = 'Returns Summary'; r['A1'].font = Font(name='Arial', bold=True, size=14)
    entries = [
        (3, 'Entry EBITDA', "='P&L'!B6", '$#,##0'),
        (4, 'Entry EV', "=Assumptions!B4*B3", '$#,##0'),
        (5, 'Entry Net Debt', "=Debt!B4", '$#,##0'),
        (6, 'Entry Equity', "=B4-B5", '$#,##0'),
        (8, 'Exit EBITDA', "='P&L'!G6", '$#,##0'),
        (9, 'Exit EV', "=Assumptions!B5*B8", '$#,##0'),
        (10, 'Exit Net Debt', "=Debt!G5", '$#,##0'),
        (11, 'Exit Equity', "=B9-B10", '$#,##0'),
        (13, 'MOIC', "=B11/B6", '0.00x'),
        (14, 'IRR', "=(B11/B6)^(1/Assumptions!B7)-1", '0.0%'),
    ]
    for row, label, formula, fmt in entries:
        r.cell(row, 1, label)
        cell = r.cell(row, 2, formula)
        cell.number_format = fmt
        cell.font = GREEN if 'Assumptions!' in formula or "'P&L'!" in formula or 'Debt!' in formula else BLACK
        if label in ('MOIC', 'IRR'):
            cell.font = Font(name='Arial', bold=True, size=12)
            r.cell(row, 1).font = BOLD
            cell.fill = YEL
        r.cell(row, 1).border = BORDER; cell.border = BORDER

    wb.save(path)

if __name__ == '__main__':
    build('/home/claude/tw/atlas_v1_base.xlsx')
    print('base built')
