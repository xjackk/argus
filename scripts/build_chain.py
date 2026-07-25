import shutil, json, subprocess, os
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

OUT = '/home/claude/chain'
os.makedirs(OUT, exist_ok=True)
BASE = '/mnt/user-data/outputs/excel-vcs-context/test-workbooks/atlas_v1_base.xlsx'
BLUE = Font(name='Arial', color='0000FF')

# Linear chain: each commit derived from the PREVIOUS file.
# (author, filename, message, mutate_fn)
def c01(wb): pass  # base

def c02(wb):
    # Deal team pushes a more aggressive growth case after mgmt meeting
    wb['Assumptions']['B11'] = 0.095

def c03(wb):
    # Modeling tightens margin ramp — exit margin was too optimistic
    wb['Assumptions']['B13'] = 0.235

def c04(wb):
    # Credit update: repricing pushes interest up, paydown accelerates
    wb['Assumptions']['B16'] = 0.0825
    wb['Assumptions']['B17'] = 0.125

def c05(wb):
    # Structural: analyst inserts a Stock-Based Comp line into P&L
    p = wb['P&L']
    p.insert_rows(7)
    p['A7'] = 'Less: Stock-Based Comp'
    for yr in range(6):
        col = get_column_letter(2 + yr)
        p[f'{col}7'] = f"=-{col}4*0.015"
        p[f'{col}7'].number_format = '$#,##0;($#,##0);-'

def c06(wb):
    # VP marks exit multiple down to reflect comp set — the headline change
    wb['Assumptions']['B5'] = 9.5

def c07(wb):
    # ⚠️ Someone overwrites the Exit EV formula with a hardcoded number
    wb['Returns']['B9'] = 2100

CHAIN = [
    ('atlas_c01_initial.xlsx',        'J. Killilea',        'Initial model shared with deal team',                c01),
    ('atlas_c02_growth_case.xlsx',    'M. Rivera',          'Revenue growth to 9.5% per mgmt meeting',            c02),
    ('atlas_c03_margin_tighten.xlsx', 'A. Chen (Modeling)', 'Tightened exit EBITDA margin to 23.5%',              c03),
    ('atlas_c04_debt_repricing.xlsx', 'A. Chen (Modeling)', 'Debt repricing: interest 8.25%, paydown 12.5%',      c04),
    ('atlas_c05_add_sbc_line.xlsx',   'M. Rivera',          'Added stock-based comp line to P&L',                 c05),
    ('atlas_c06_exit_multiple.xlsx',  'S. Patel (VP)',      'Marked exit multiple down to 9.5x per comp set',     c06),
    ('atlas_c07_hardcode_flag.xlsx',  'M. Rivera',          'Manual Exit EV override pending diligence',          c07),
]

prev = BASE
commits = []
for i, (fname, author, msg, fn) in enumerate(CHAIN):
    dst = os.path.join(OUT, fname)
    shutil.copy(prev, dst)
    wb = load_workbook(dst)
    fn(wb)
    wb.save(dst)
    subprocess.run(['python3', '/mnt/skills/public/xlsx/scripts/recalc.py', dst, '60'],
                   capture_output=True)
    commits.append({'file': fname, 'author': author, 'message': msg})
    prev = dst
    print(f'  {i+1}. {fname:<32} {author}')

print('\nchain built')
