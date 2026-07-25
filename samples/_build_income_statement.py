#!/usr/bin/env python3
"""Build LARGE, richly-formatted Income Statement demo workbooks (v1/v2).

Assumptions sheet (blue input cells) drives a P&L sheet across 8 periods
(FY20-FY27) with 18 line items, every value a cross-sheet / cross-cell formula.
v1 -> v2 changes ONE input (Revenue Growth %), which ripples through dozens of
computed cells. Labels live in column A (the engine reads row labels from A).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CUR = '$#,##0'
PCT = '0.0%'

INPUT_FILL = PatternFill(fill_type='solid', fgColor='DCE6F1')      # light blue
INPUT_FONT = Font(color='1F4E78', bold=True)                        # dark blue
TITLE_FONT = Font(bold=True, size=13, color='1F4E78')
HDR_FONT = Font(bold=True, color='FFFFFF')
HDR_FILL = PatternFill(fill_type='solid', fgColor='1F4E78')
LABEL_FONT = Font(bold=False)
BOLD = Font(bold=True)
TOP_BORDER = Border(top=Side(style='thin', color='808080'))

PERIODS = ['FY20', 'FY21', 'FY22', 'FY23', 'FY24', 'FY25', 'FY26', 'FY27']
FIRST_COL = 2  # column B


def build(growth, path):
    wb = Workbook()

    # ---------------- Assumptions sheet ----------------
    a = wb.active
    a.title = 'Assumptions'
    a['A1'] = 'Assumptions & Drivers'
    a['A1'].font = TITLE_FONT
    a['A2'] = '(blue cells are authored inputs)'
    a['A2'].font = Font(italic=True, size=9, color='808080')

    # (label, value, number_format) starting at row 3 -> B3 is first input
    inputs = [
        ('Base Revenue (FY20, $000)', 42000, CUR),   # B3
        ('Revenue Growth %',          growth, PCT),   # B4  <-- THE CHANGE
        ('Gross Margin %',            0.62,  PCT),    # B5
        ('SG&A % of Revenue',         0.14,  PCT),    # B6
        ('R&D % of Revenue',          0.09,  PCT),    # B7
        ('Marketing % of Revenue',    0.07,  PCT),    # B8
        ('G&A % of Revenue',          0.05,  PCT),    # B9
        ('D&A % of Revenue',          0.04,  PCT),    # B10
        ('Interest Rate on Debt',     0.06,  PCT),    # B11
        ('Debt Balance ($000)',       20000, CUR),    # B12
        ('Tax Rate',                  0.24,  PCT),    # B13
    ]
    ROW = {}
    r = 3
    for label, val, fmt in inputs:
        a.cell(row=r, column=1, value=label).font = LABEL_FONT
        c = a.cell(row=r, column=2, value=val)
        c.number_format = fmt
        c.fill = INPUT_FILL
        c.font = INPUT_FONT
        ROW[label] = r
        r += 1
    a.column_dimensions['A'].width = 30
    a.column_dimensions['B'].width = 14

    # convenient absolute refs into Assumptions
    def A(label):
        return f"Assumptions!$B${ROW[label]}"

    # ---------------- P&L sheet ----------------
    p = wb.create_sheet('P&L')
    p['A1'] = 'Income Statement (USD $000)'
    p['A1'].font = TITLE_FONT
    p['A2'] = '($ in thousands)'
    p['A2'].font = Font(italic=True, size=9, color='808080')
    for i, per in enumerate(PERIODS):
        cell = p.cell(row=2, column=FIRST_COL + i, value=per)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='right')

    # line item rows: (key, label, format, is_total_row)
    layout = [
        ('rev',   'Revenue',                    CUR, False),
        ('cogs',  'Cost of Goods Sold',         CUR, False),
        ('gp',    'Gross Profit',               CUR, True),
        ('sga',   'SG&A',                       CUR, False),
        ('rd',    'Research & Development',     CUR, False),
        ('mkt',   'Marketing',                  CUR, False),
        ('ga',    'General & Administrative',   CUR, False),
        ('opex',  'Total Operating Expenses',   CUR, True),
        ('ebitda','EBITDA',                     CUR, True),
        ('da',    'Depreciation & Amortization',CUR, False),
        ('ebit',  'Operating Income (EBIT)',    CUR, True),
        ('int',   'Interest Expense',           CUR, False),
        ('pre',   'Pretax Income',              CUR, True),
        ('tax',   'Income Tax',                 CUR, False),
        ('ni',    'Net Income',                 CUR, True),
        ('gm',    'Gross Margin %',             PCT, False),
        ('em',    'EBITDA Margin %',            PCT, False),
        ('nm',    'Net Margin %',               PCT, False),
    ]
    R = {}
    r = 3
    for key, label, fmt, is_total in layout:
        lc = p.cell(row=r, column=1, value=label)
        lc.font = BOLD if is_total else LABEL_FONT
        R[key] = r
        r += 1

    # write formulas per period column
    for i, per in enumerate(PERIODS):
        col = get_column_letter(FIRST_COL + i)
        prev = get_column_letter(FIRST_COL + i - 1) if i > 0 else None

        def put(key, formula, fmt):
            c = p.cell(row=R[key], column=FIRST_COL + i, value=formula)
            c.number_format = fmt
            if layout_is_total(key):
                c.font = BOLD
                c.border = TOP_BORDER

        # Revenue: base for FY20, grow off prior column thereafter
        if i == 0:
            put('rev', f"={A('Base Revenue (FY20, $000)')}", CUR)
        else:
            put('rev', f"={prev}{R['rev']}*(1+{A('Revenue Growth %')})", CUR)

        put('cogs', f"={col}{R['rev']}*(1-{A('Gross Margin %')})", CUR)
        put('gp',   f"={col}{R['rev']}-{col}{R['cogs']}", CUR)
        put('sga',  f"={col}{R['rev']}*{A('SG&A % of Revenue')}", CUR)
        put('rd',   f"={col}{R['rev']}*{A('R&D % of Revenue')}", CUR)
        put('mkt',  f"={col}{R['rev']}*{A('Marketing % of Revenue')}", CUR)
        put('ga',   f"={col}{R['rev']}*{A('G&A % of Revenue')}", CUR)
        put('opex', f"=SUM({col}{R['sga']}:{col}{R['ga']})", CUR)
        put('ebitda', f"={col}{R['gp']}-{col}{R['opex']}", CUR)
        put('da',   f"={col}{R['rev']}*{A('D&A % of Revenue')}", CUR)
        put('ebit', f"={col}{R['ebitda']}-{col}{R['da']}", CUR)
        put('int',  f"={A('Debt Balance ($000)')}*{A('Interest Rate on Debt')}", CUR)
        put('pre',  f"={col}{R['ebit']}-{col}{R['int']}", CUR)
        put('tax',  f"={col}{R['pre']}*{A('Tax Rate')}", CUR)
        put('ni',   f"={col}{R['pre']}-{col}{R['tax']}", CUR)
        put('gm',   f"={col}{R['gp']}/{col}{R['rev']}", PCT)
        put('em',   f"={col}{R['ebitda']}/{col}{R['rev']}", PCT)
        put('nm',   f"={col}{R['ni']}/{col}{R['rev']}", PCT)

    p.column_dimensions['A'].width = 28
    for i in range(len(PERIODS)):
        p.column_dimensions[get_column_letter(FIRST_COL + i)].width = 12

    wb.save(path)
    print('wrote', path)


_TOTAL_KEYS = {'gp', 'opex', 'ebitda', 'ebit', 'pre', 'ni'}
def layout_is_total(key):
    return key in _TOTAL_KEYS


if __name__ == '__main__':
    build(0.10, 'income_statement_v1.xlsx')
    build(0.16, 'income_statement_v2.xlsx')
