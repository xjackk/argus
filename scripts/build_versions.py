import shutil
from openpyxl import load_workbook

BASE = '/home/claude/tw/atlas_v1_base.xlsx'

# ---- V2: single authored input change -> pure cascade test ----
# Lower exit multiple 10.5x -> 9.5x. One input edit; everything downstream recomputes.
shutil.copy(BASE, '/home/claude/tw/atlas_v2_exit_multiple.xlsx')
wb = load_workbook('/home/claude/tw/atlas_v2_exit_multiple.xlsx')
wb['Assumptions']['B5'] = 9.5
wb.save('/home/claude/tw/atlas_v2_exit_multiple.xlsx')

# ---- V3: multiple authored inputs across sheets ----
# Growth 8%->6.5%, entry margin 22%->20%, interest 7.5%->8.25%. Bigger blast radius.
shutil.copy(BASE, '/home/claude/tw/atlas_v3_downside.xlsx')
wb = load_workbook('/home/claude/tw/atlas_v3_downside.xlsx')
wb['Assumptions']['B11'] = 0.065
wb['Assumptions']['B12'] = 0.20
wb['Assumptions']['B16'] = 0.0825
wb.save('/home/claude/tw/atlas_v3_downside.xlsx')

# ---- V4: STRUCTURAL change — inserted row on P&L ----
# Insert a "Stock-Based Comp" line. This shifts rows below it; a naive positional
# diff will flag the whole lower half of the sheet as changed. THE alignment test.
shutil.copy(BASE, '/home/claude/tw/atlas_v4_inserted_row.xlsx')
wb = load_workbook('/home/claude/tw/atlas_v4_inserted_row.xlsx')
p = wb['P&L']
p.insert_rows(7)  # insert above D&A (row 7), pushing EBIT/interest/etc down by 1
p['A7'] = 'Less: Stock-Based Comp'
for yr in range(6):
    from openpyxl.utils import get_column_letter
    col = get_column_letter(2+yr)
    p[f'{col}7'] = f"=-{col}4*0.015"  # 1.5% of revenue
    p[f'{col}7'].number_format = '$#,##0;($#,##0);-'
# NOTE: downstream formulas that referenced row 7 (old D&A) now need to still work;
# openpyxl does NOT auto-adjust formula references on insert_rows, so we deliberately
# leave this as-is to mimic a REAL messy human edit where someone inserts a row and
# the model partially breaks — good adversarial input for the diff engine + recalc check.
wb.save('/home/claude/tw/atlas_v4_inserted_row.xlsx')

# ---- V5: the "smells wrong" test — hardcode replacing a formula ----
# Someone overwrites the exit-multiple-driven Exit EV formula with a hardcoded number.
# This is exactly the anomaly the AI-flagging feature should catch (formula -> constant).
shutil.copy(BASE, '/home/claude/tw/atlas_v5_hardcode_override.xlsx')
wb = load_workbook('/home/claude/tw/atlas_v5_hardcode_override.xlsx')
wb['Returns']['B9'] = 2100  # was "=Assumptions!B5*B8"; now a hardcoded override
wb.save('/home/claude/tw/atlas_v5_hardcode_override.xlsx')

print('version pairs written')
